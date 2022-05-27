import openpyxl
import matplotlib.pyplot as plt

arquivo_arenado = openpyxl.load_workbook('./gamelogs/gamelogs_arenado.xlsx')
arquivo_goldy = openpyxl.load_workbook('./gamelogs/gamelogs_goldy.xlsx')
arquivo_time = openpyxl.load_workbook('./gamelogs/gamelogs_time.xlsx')

gamelogs_arenado = arquivo_arenado['Worksheet']
gamelogs_goldy = arquivo_goldy['Worksheet']
gamelogs_time = arquivo_time['Worksheet']
obp_arenado = []
datas_arenado = []
obp_goldy = []
datas_goldy = []

obp_time = []
datas_time = []


for i in range(2, gamelogs_arenado.max_row):
    obp_arenado.append(gamelogs_arenado[f'AE{i}'].value)
    datas_arenado.append(gamelogs_arenado[f'C{i}'].value)
    
for i in range(2, gamelogs_goldy.max_row):
    obp_goldy.append(gamelogs_goldy[f'AE{i}'].value)
    datas_goldy.append(gamelogs_goldy[f'C{i}'].value)
    
for i in range(2, gamelogs_time.max_row):
    obp_time.append(gamelogs_time[f'AB{i}'].value)
    datas_time.append(gamelogs_time[f'B{i}'].value)

plt.plot(datas_arenado, obp_arenado, marker = 'o', label = 'Nolan Arenado')
plt.plot(datas_goldy, obp_goldy, marker = 'o', label = 'Paul Goldschmidt')
plt.plot(datas_time, obp_time, marker = 'o', label = 'Time')
plt.legend()
plt.title('OPS de Nolan Arenado e Paul Goldschmidt pela temporada')
plt.xlabel('Dia')
plt.ylabel('OPS')
plt.show()